#!/usr/bin/env python3
"""
World Cup 2026 tracker auto-updater for Steve.

Pulls live results from ESPN's (undocumented, free, no-key) FIFA World Cup
scoreboard endpoint and rebuilds an Excel tracker for your four teams:
Morocco, Canada, Jordan, Iraq.

Usage:
    python update_wc.py                 # fetch live data, write wc_tracker.xlsx
    python update_wc.py --out C:\\path\\wc.xlsx
    python update_wc.py --fixture sample.json   # offline test from a saved JSON file

Dependencies: openpyxl  (pip install openpyxl). Everything else is stdlib.
"""

import argparse, datetime as dt, json, os, sys, unicodedata, urllib.request

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---- config ---------------------------------------------------------------
# Single source of truth for the whole pool: owner -> 4 teams. Each team is a
# list of name VARIANTS (first = display name). Matching is accent/punctuation-
# insensitive via norm(), so only genuinely DIFFERENT spellings need listing
# (e.g. ESPN "Ivory Coast" vs draft "Cote d'Ivoire", ESPN "Morocco" vs the
# draft sheet's "Monaco" typo). Owner names are taken from the draft sheet
# (World Cup 2026.xlsx). ME is which owner is Steve.
ME = "STEVE S"  # Steve Simon. NOTE: the draft sheet still labels this row "JOHN C"
                # (John Vinson) — John dropped out and Steve took over his 4 teams
                # (led by Canada), per Doug's 2026-06-10 email. Owner shown as STEVE S.
OWNERS = {
    "DOUG":   [["Ecuador"], ["Ghana"], ["Senegal"], ["Argentina"]],
    "JEFF":   [["Türkiye", "Turkiye", "Turkey"], ["Sweden"],
               ["Curaçao", "Curacao"], ["United States", "USA"]],
    "VICKI":  [["Panama"], ["Ivory Coast", "Côte d'Ivoire", "Cote d'Ivoire"],
               ["Algeria"], ["England"]],
    "VIC":    [["Colombia", "Columbia"], ["Uzbekistan"],
               ["Cape Verde", "Cabo Verde"], ["Belgium"]],
    "AMY":    [["New Zealand"], ["Haiti"], ["Congo DR", "DR Congo"], ["France"]],
    "MARK":   [["Uruguay"], ["Austria"], ["Egypt"], ["Germany"]],
    "BOB M":  [["Norway"], ["Saudi Arabia"], ["Japan"], ["Spain"]],
    "STEVE S": [["Morocco", "Monaco"], ["Canada"], ["Jordan"], ["Iraq"]],  # was JOHN C
    "SCOTT":  [["Scotland"], ["Iran", "IR Iran"], ["Croatia"], ["Brazil"]],
    "SOHAIL": [["Tunisia"], ["Australia"],
               ["Bosnia-Herzegovina", "Bosnia and Herzegovina", "Bosnia Herzegovina"],
               ["Mexico"]],
    "JOE S":  [["Switzerland"], ["South Korea", "Korea Republic", "S. Korea", "Korea"],
               ["Paraguay"], ["Netherlands"]],
    "ED":     [["Czechia"], ["Qatar"], ["South Africa", "S. Africa"], ["Portugal"]],
}
MY_TEAMS = [variants[0] for variants in OWNERS[ME]]   # display names of Steve's teams

ENDPOINT = ("https://site.api.espn.com/apis/site/v2/sports/soccer/"
            "fifa.world/scoreboard?limit=400&dates=20260611-20260719")
ET_OFFSET = dt.timedelta(hours=-4)   # Eastern Daylight Time (fixed; DST active all tournament)


def norm(s):
    """Accent/punctuation/space-insensitive key for name matching."""
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return "".join(c for c in s.lower() if c.isalnum())


# normalized variant -> (owner, display_name); built once from OWNERS.
TEAM_LUT = {}
for _owner, _teams in OWNERS.items():
    for _variants in _teams:
        for _v in _variants:
            TEAM_LUT[norm(_v)] = (_owner, _variants[0])


def resolve(name):
    """ESPN team displayName -> (owner, display_name), or None if not a pool team
    (e.g. knockout placeholders like 'Group C Winner', 'Round of 32 3 Winner')."""
    return TEAM_LUT.get(norm(name))

DEFAULT_FILENAME = "World Cup 2026 Tracker.xlsx"


def default_output_path():
    home = os.path.expanduser("~")
    desktop = os.path.join(home, "Desktop")
    if os.path.isdir(desktop):
        return os.path.join(desktop, DEFAULT_FILENAME)
    onedrive = os.environ.get("OneDrive") or os.environ.get("ONEDRIVE")
    if onedrive:
        od_desktop = os.path.join(onedrive, "Desktop")
        if os.path.isdir(od_desktop):
            return os.path.join(od_desktop, DEFAULT_FILENAME)
    return os.path.join(os.getcwd(), DEFAULT_FILENAME)

# ---- data fetch -----------------------------------------------------------
def fetch_events(fixture=None):
    """Return the list of event dicts from ESPN (or a local fixture file)."""
    if fixture:
        with open(fixture, "r", encoding="utf-8") as f:
            return json.load(f).get("events", [])
    req = urllib.request.Request(
        ENDPOINT, headers={"User-Agent": "Mozilla/5.0 (wc-tracker)"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8")).get("events", [])

# ---- parsing --------------------------------------------------------------
def parse_match(event, my_team):
    """Pull the fields we care about for a given event from my_team's view."""
    comp = event["competitions"][0]
    state = comp["status"]["type"]["state"]            # pre | in | post
    completed = comp["status"]["type"].get("completed", False)
    mine = opp = None
    for c in comp["competitors"]:
        r = resolve(c["team"]["displayName"])
        if r and r[1] == my_team:
            mine = c
        else:
            opp = c
    if mine is None or opp is None:
        return None

    # kickoff in ET
    utc = dt.datetime.fromisoformat(event["date"].replace("Z", "+00:00"))
    et = utc + ET_OFFSET
    when = et.strftime("%a %b %-d, %-I:%M %p ET") if sys.platform != "win32" \
        else et.strftime("%a %b %#d, %#I:%M %p ET")

    played = state in ("in", "post")
    gf = int(mine["score"]) if played and mine["score"] not in ("", None) else None
    ga = int(opp["score"]) if played and opp["score"] not in ("", None) else None

    if state == "pre":
        status = when
    elif state == "in":
        status = "LIVE " + comp["status"]["type"].get("shortDetail", "")
    else:
        status = "Final"

    return {
        "date_sort": et,
        "when": when,
        "round": comp.get("altGameNote", "").replace("FIFA World Cup, ", ""),
        "opponent": ("vs " if mine["homeAway"] == "home" else "at ") + opp["team"]["displayName"],
        "gf": gf, "ga": ga, "state": state, "completed": completed, "status": status,
    }

def collect(events):
    """Return {team: [match, ...]} sorted by date for each of MY_TEAMS."""
    out = {t: [] for t in MY_TEAMS}
    for ev in events:
        disps = []
        for c in ev["competitions"][0]["competitors"]:
            r = resolve(c["team"]["displayName"])
            disps.append(r[1] if r else None)
        for t in MY_TEAMS:
            if t in disps:
                m = parse_match(ev, t)
                if m:
                    out[t].append(m)
    for t in out:
        out[t].sort(key=lambda m: m["date_sort"])
    return out

# ---- pool aggregation -----------------------------------------------------
def pool_table(events):
    """Aggregate every pool team across all events. Returns a list of dicts
    sorted by the standard soccer order: Points -> goal diff -> goals for.
    Live ('in') games count provisionally, same as the My Teams tab."""
    # seed all 48 so teams with no game yet still show (P=0)
    stats = {}
    for owner, teams in OWNERS.items():
        for variants in teams:
            stats[variants[0]] = {
                "team": variants[0], "owner": owner,
                "p": 0, "w": 0, "d": 0, "l": 0, "gf": 0, "ga": 0,
            }
    for ev in events:
        comp = ev["competitions"][0]
        state = comp["status"]["type"]["state"]
        if state not in ("in", "post"):
            continue
        rows = []
        for c in comp["competitors"]:
            r = resolve(c["team"]["displayName"])
            if not r or c["score"] in ("", None):
                rows = []
                break
            rows.append((r[1], int(c["score"])))
        if len(rows) != 2:
            continue
        (ta, sa), (tb, sb) = rows
        for team, gf, ga in ((ta, sa, sb), (tb, sb, sa)):
            s = stats[team]
            s["p"] += 1; s["gf"] += gf; s["ga"] += ga
            if gf > ga:   s["w"] += 1
            elif gf < ga: s["l"] += 1
            else:         s["d"] += 1
    for s in stats.values():
        s["pts"] = 3 * s["w"] + s["d"]
        s["gd"] = s["gf"] - s["ga"]
    return sorted(stats.values(),
                  key=lambda s: (-s["pts"], -s["gd"], -s["gf"], s["team"]))

# ---- workbook -------------------------------------------------------------
FONT = "Arial"
NAVY, MID, GREY, YELLOW, GREEN, LIVE = "1F3864", "2E5496", "D9D9D9", "FFF2CC", "C6EFCE", "FCE4D6"
TEAM_CLR = {"Morocco": "7E2D2D", "Canada": "8A2A2A", "Jordan": "5A3D2B", "Iraq": "3D5A2B"}
GROUP = {"Morocco": "C", "Canada": "B", "Jordan": "J", "Iraq": "I"}
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def _set(cell, val=None, bold=False, size=10, color="000000", fill=None,
         align="left", wrap=False):
    if val is not None:
        cell.value = val
    cell.font = Font(name=FONT, bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = PatternFill("solid", start_color=fill)

def build_xlsx(data, pool, path):
    wb = Workbook(); ws = wb.active; ws.title = "My Teams"
    for col, w in {"A": 22, "B": 22, "C": 7, "D": 7, "E": 8, "F": 6, "G": 22}.items():
        ws.column_dimensions[col].width = w

    stamp = dt.datetime.now(dt.timezone.utc).replace(tzinfo=None) + ET_OFFSET
    ws.merge_cells("A1:G1")
    _set(ws["A1"], "STEVE'S WORLD CUP 2026 TRACKER", bold=True, size=15,
         color="FFFFFF", fill=NAVY, align="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:G2")
    _set(ws["A2"], f"Auto-updated from ESPN  |  {stamp:%a %b %d, %I:%M %p ET}  |  "
                   "Pool format: winner-take-all (only the champion matters)",
         size=9, color="404040", align="center")

    # leave rows 4-10 for summary, build detail from row 12
    row = 12
    sub = {}
    for t in MY_TEAMS:
        ws.merge_cells(f"A{row}:G{row}")
        _set(ws[f"A{row}"], f"{t.upper()}  —  Group {GROUP[t]}", bold=True, size=11,
             color="FFFFFF", fill=TEAM_CLR[t])
        row += 1
        for i, h in enumerate(["Round", "Opponent", "GF", "GA", "Result", "Pts", "Status"]):
            c = ws.cell(row=row, column=1 + i)
            _set(c, h, bold=True, size=9, color="FFFFFF", fill=MID,
                 align="left" if i in (0, 1, 6) else "center")
            c.border = BORDER
        row += 1
        first = row
        for m in data[t]:
            _set(ws.cell(row=row, column=1), m["round"] or "Group", size=9)
            _set(ws.cell(row=row, column=2), m["opponent"], size=10)
            gfc, gac = ws.cell(row=row, column=3), ws.cell(row=row, column=4)
            if m["gf"] is not None:
                _set(gfc, m["gf"], align="center"); _set(gac, m["ga"], align="center")
            else:
                _set(gfc, None, align="center", fill=YELLOW)
                _set(gac, None, align="center", fill=YELLOW)
            rc = ws.cell(row=row, column=5)
            rc.value = (f'=IF(OR(C{row}="",D{row}=""),"",'
                        f'IF(C{row}>D{row},"W",IF(C{row}<D{row},"L","D")))')
            _set(rc, bold=True, align="center")
            pc = ws.cell(row=row, column=6)
            pc.value = f'=IF(E{row}="","",IF(E{row}="W",3,IF(E{row}="D",1,0)))'
            _set(pc, align="center")
            fill = LIVE if m["state"] == "in" else None
            _set(ws.cell(row=row, column=7), m["status"], size=9,
                 color="C00000" if m["state"] == "in" else "808080", fill=fill,
                 bold=(m["state"] == "in"))
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = BORDER
            row += 1
        last = row - 1
        # subtotal
        _set(ws.cell(row=row, column=2), "Subtotal", bold=True, fill=GREY, align="right")
        gf_c, ga_c = ws.cell(row=row, column=3), ws.cell(row=row, column=4)
        gf_c.value = f"=SUM(C{first}:C{last})" if last >= first else 0
        ga_c.value = f"=SUM(D{first}:D{last})" if last >= first else 0
        _set(gf_c, bold=True, align="center", fill=GREY)
        _set(ga_c, bold=True, align="center", fill=GREY)
        rec = ws.cell(row=row, column=5)
        rec.value = (f'=COUNTIF(E{first}:E{last},"W")&"-"&COUNTIF(E{first}:E{last},"D")'
                     f'&"-"&COUNTIF(E{first}:E{last},"L")') if last >= first else "0-0-0"
        _set(rec, bold=True, align="center", size=9, fill=GREY)
        ptc = ws.cell(row=row, column=6)
        ptc.value = (f'=3*COUNTIF(E{first}:E{last},"W")+COUNTIF(E{first}:E{last},"D")'
                     ) if last >= first else 0
        _set(ptc, bold=True, align="center", size=11, fill=GREY)
        _set(ws.cell(row=row, column=7), None, fill=GREY)
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER
        sub[t] = {"pts": f"F{row}", "rec_first": first, "rec_last": last}
        row += 2

    # summary table at top
    sr = 4
    ws.merge_cells(f"A{sr}:G{sr}")
    _set(ws[f"A{sr}"], "SUMMARY  —  YOUR TEAMS", bold=True, size=12, color="FFFFFF", fill=MID)
    sr += 1
    for i, h in enumerate(["Team", "Group", "Played", "W-D-L", "Pts", "Title shot", ""]):
        c = ws.cell(row=sr, column=1 + i)
        _set(c, h, bold=True, color="FFFFFF", fill=NAVY,
             align="left" if i in (0, 1, 5) else "center")
        c.border = BORDER
    sr += 1
    note = {"Morocco": "LIVE — dark horse", "Canada": "≈ none", "Jordan": "≈ none", "Iraq": "≈ none"}
    for t in MY_TEAMS:
        s = sub[t]; f, l = s["rec_first"], s["rec_last"]
        _set(ws.cell(row=sr, column=1), t, bold=True)
        _set(ws.cell(row=sr, column=2), f"Group {GROUP[t]}")
        pl = ws.cell(row=sr, column=3)
        pl.value = (f'=COUNTIF(E{f}:E{l},"W")+COUNTIF(E{f}:E{l},"D")+COUNTIF(E{f}:E{l},"L")'
                    ) if l >= f else 0
        _set(pl, align="center")
        wdl = ws.cell(row=sr, column=4)
        wdl.value = (f'=COUNTIF(E{f}:E{l},"W")&"-"&COUNTIF(E{f}:E{l},"D")&"-"&'
                     f'COUNTIF(E{f}:E{l},"L")') if l >= f else "0-0-0"
        _set(wdl, align="center")
        pt = ws.cell(row=sr, column=5); pt.value = f"={s['pts']}"
        _set(pt, bold=True, align="center", size=11)
        _set(ws.cell(row=sr, column=6), note[t], size=9,
             color="2E7D32" if t == "Morocco" else "808080",
             bold=(t == "Morocco"))
        for col in range(1, 7):
            ws.cell(row=sr, column=col).border = BORDER
        sr += 1
    _set(ws.cell(row=sr, column=1), "Pot: $240  |  Your shot rides entirely on Morocco",
         bold=True, size=10, fill=GREEN)
    ws.merge_cells(f"A{sr}:G{sr}")
    _set(ws[f"A{sr}"], "Pot: $240  |  Winner-take-all  |  Your shot rides entirely on Morocco's run",
         bold=True, size=10, fill=GREEN, align="center")
    for col in range(1, 8):
        ws.cell(row=sr, column=col).fill = PatternFill("solid", start_color=GREEN)

    ws.freeze_panes = "A12"

    build_pool(wb, pool)
    wb.save(path)

# ---- pool tab -------------------------------------------------------------
def build_pool(wb, pool):
    """Second tab: all 48 teams with owner, points, and tie-breaker (GD),
    sorted Points -> GD -> GF. Steve's teams highlighted."""
    ws = wb.create_sheet("Pool")
    widths = {"A": 5, "B": 22, "C": 9, "D": 7, "E": 9, "F": 6, "G": 6,
              "H": 11, "I": 6}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:I1")
    _set(ws["A1"], "POOL STANDINGS  —  ALL 48 TEAMS", bold=True, size=14,
         color="FFFFFF", fill=NAVY, align="center")
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A2:I2")
    _set(ws["A2"], "Sorted by Points, then Tie-Breaker (goal difference), then "
                   "goals for  |  Pts/Tiebreak are for tracking only — pool is "
                   "winner-take-all  |  your teams shaded",
         size=9, color="404040", align="center")

    headers = ["#", "Team", "Owner", "Played", "W-D-L", "GF", "GA",
               "Tiebreak (GD)", "Pts"]
    hrow = 4
    for i, h in enumerate(headers):
        c = ws.cell(row=hrow, column=1 + i)
        _set(c, h, bold=True, size=9, color="FFFFFF", fill=MID,
             align="left" if i in (1, 2) else "center")
        c.border = BORDER

    row = hrow + 1
    for rank, s in enumerate(pool, start=1):
        mine = s["owner"] == ME
        shade = YELLOW if mine else (GREY if rank % 2 == 0 else None)
        _set(ws.cell(row=row, column=1), rank, align="center", size=9, fill=shade)
        _set(ws.cell(row=row, column=2), s["team"], bold=mine, fill=shade)
        _set(ws.cell(row=row, column=3), s["owner"],
             bold=mine, size=9, fill=shade,
             color="1F3864" if mine else "404040")
        _set(ws.cell(row=row, column=4), s["p"], align="center", fill=shade)
        _set(ws.cell(row=row, column=5),
             f'{s["w"]}-{s["d"]}-{s["l"]}', align="center", size=9, fill=shade)
        _set(ws.cell(row=row, column=6), s["gf"], align="center", fill=shade)
        _set(ws.cell(row=row, column=7), s["ga"], align="center", fill=shade)
        gd = ws.cell(row=row, column=8)
        _set(gd, f'{s["gd"]:+d}' if s["p"] else 0, align="center", fill=shade)
        _set(ws.cell(row=row, column=9), s["pts"], bold=True, align="center",
             size=11, fill=shade)
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    ws.freeze_panes = "A5"

# ---- live web page (Android-friendly) -------------------------------------
# Self-contained HTML for GitHub Pages. The page fetches ESPN directly in the
# browser (ESPN sends Access-Control-Allow-Origin: *), recomputes the standings
# with the SAME OWNERS map embedded below, and auto-refreshes every 60s. No
# server, no scheduled job — every visitor's phone pulls fresh scores on load.
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>World Cup 2026 Pool</title>
<!-- inline SVG favicon (soccer ball) — keeps the page a single self-contained file -->
<link rel="icon" href="data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'%3E%3Ctext y='.9em' font-size='90'%3E%E2%9A%BD%3C/text%3E%3C/svg%3E">
<link rel="apple-touch-icon" href="data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'%3E%3Crect width='100' height='100' fill='%231f3864'/%3E%3Ctext y='.9em' x='5' font-size='80'%3E%E2%9A%BD%3C/text%3E%3C/svg%3E">
<meta name="theme-color" content="#1f3864">
<style>
  *{box-sizing:border-box}
  body{margin:0;font-family:-apple-system,"Segoe UI",Roboto,Arial,sans-serif;
       background:#0f1626;color:#e8edf6;-webkit-text-size-adjust:100%}
  header{background:#1f3864;padding:14px 12px;text-align:center}
  header h1{margin:0;font-size:20px;letter-spacing:.5px}
  header .sub{font-size:12px;opacity:.85;margin-top:3px}
  .nav{display:flex;gap:16px;justify-content:center;background:#16203a;
       padding:9px 12px;border-bottom:1px solid #1c2740;font-size:13px}
  .nav a{color:#9fb0cc;text-decoration:none;font-weight:600}
  .nav a.active{color:#ffd86b}
  .status{font-size:12px;text-align:center;padding:8px;color:#9fb0cc}
  .status .live{color:#ff6b6b;font-weight:700}
  .controls{display:flex;align-items:center;justify-content:center;gap:8px;
            flex-wrap:wrap;padding:10px 12px;background:#16203a;
            border-bottom:1px solid #1c2740}
  .controls label{font-size:13px;color:#9fb0cc}
  .controls select{background:#0f1626;color:#e8edf6;border:1px solid #2a3a5e;
                   border-radius:8px;padding:7px 10px;font-size:15px}
  .mine{padding:6px 12px 10px}
  .mine .hint{font-size:13px;color:#7d8db0}
  .mine h2{font-size:11px;margin:0 0 6px;color:#9fb0cc;font-weight:600;
           text-transform:uppercase;letter-spacing:.5px}
  .chips{display:flex;flex-wrap:wrap;gap:6px}
  .chip{background:#16203a;border:1px solid #2a3a5e;border-radius:8px;
        padding:6px 9px;font-size:13px}
  .chip b{color:#ffd86b}
  table{width:100%;border-collapse:collapse;font-size:14px}
  thead th{position:sticky;top:0;background:#16203a;color:#9fb0cc;font-size:11px;
           text-transform:uppercase;letter-spacing:.3px;padding:8px 6px;text-align:center}
  thead th:nth-child(2),thead th:nth-child(3){text-align:left}
  tbody td{padding:8px 6px;text-align:center;border-bottom:1px solid #1c2740}
  tbody td:nth-child(2),tbody td:nth-child(3){text-align:left}
  tbody tr:nth-child(even){background:#121c33}
  tbody tr.mine-row{background:#fff2cc;color:#1a1a1a}
  tbody tr.mine-row td{border-bottom-color:#e6d9a8;font-weight:600}
  .team{font-weight:600}
  .pts{font-weight:700;font-size:15px}
  .ssub{font-size:11px;color:#7d8db0;font-weight:400;margin-top:2px;line-height:1.3}
  .ssub.dead{color:#6b6b6b}
  /* in the table the next-match detail is collapsed: hover (desktop) or tap (mobile) reveals it */
  #rows .ssub{display:none}
  #rows tr:hover .ssub,#rows tr.expanded .ssub{display:block}
  #rows .team{cursor:pointer}
  .caret{display:inline-block;margin-left:5px;color:#6b7a99;font-weight:700;font-size:12px;
         transition:transform .15s}
  #rows tr:hover .caret,#rows tr.expanded .caret{transform:rotate(90deg);color:#9fb0cc}
  .bdg{display:inline-block;margin-left:6px;font-size:10px;font-weight:700;
       padding:1px 6px;border-radius:6px;vertical-align:middle;letter-spacing:.3px}
  .bdg.live{background:#1f7a4d;color:#dffff0}
  .bdg.out{background:#3a3a3a;color:#b6b6b6}
  .bdg.champ{background:#ffd86b;color:#1a1a1a}
  tbody tr.out{opacity:.55}
  tbody tr.out .team .tname{text-decoration:line-through;text-decoration-color:#7a7a7a}
  tbody tr.mine-row.out{opacity:.7}
  .chip{display:inline-block}
  .chip .ssub{margin-top:3px}
  footer{font-size:11px;text-align:center;color:#6b7a99;padding:14px 12px 24px;line-height:1.5}
</style>
</head>
<body>
<header>
  <h1>⚽ WORLD CUP 2026 POOL</h1>
  <div class="sub">Winner-take-all · $240 pot · only the champion's owner wins</div>
</header>
<nav class="nav"><a class="active" href="index.html">Standings</a><a href="bracket.html">Bracket →</a></nav>
<div class="controls">
  <label for="me">⭐ Highlight my teams:</label>
  <select id="me"><option value="">— pick your name —</option></select>
</div>
<div class="status" id="status">Loading live scores…</div>
<section class="mine" id="mine"></section>
<table>
  <thead><tr><th>#</th><th>Team</th><th>Owner</th><th>P</th><th>W-D-L</th><th>GD</th><th>Pts</th></tr></thead>
  <tbody id="rows"></tbody>
</table>
<footer>
  Live from ESPN (unofficial) · auto-refreshes every 60s · your teams in gold<br>
  Points &amp; goal difference are for bragging rights — the pool is winner-take-all.
</footer>
<script>
const OWNERS = %%OWNERS%%;
const ENDPOINT = "%%ENDPOINT%%";

function norm(s){
  return (s||"").normalize("NFKD").replace(/[̀-ͯ]/g,"")
                .toLowerCase().replace(/[^a-z0-9]/g,"");
}
const LUT = {};
for (const [owner, teams] of Object.entries(OWNERS))
  for (const variants of teams)
    for (const v of variants) LUT[norm(v)] = [owner, variants[0]];
function resolve(name){ return LUT[norm(name)] || null; }

function fmtGD(s){ return s.p ? (s.gd > 0 ? "+" : "") + s.gd : "0"; }

// --- knockout status + next-match per team ---------------------------------
// Uses ESPN's winner/advance flags (correct even when a tie is decided on
// penalties) plus each team's earliest upcoming fixture for the "next match"
// subline. Returns {team: {badge, cls, sub, dead, champ}}.
const KO_ROUNDS = ["Round of 32","Round of 16","Quarterfinals","Semifinals","Final"];
const KO_SHORT  = {"Round of 32":"R32","Round of 16":"R16","Quarterfinals":"QF",
                   "Semifinals":"SF","Final":"Final"};
function koLevel(note){
  for (let i = 0; i < KO_ROUNDS.length; i++)
    if (note.endsWith(KO_ROUNDS[i])) return i;
  return -1;
}
function buildInfo(events, teams){
  const info = {};
  for (const t of teams) info[t] = {inKO:false, champ:false, exit:null, next:null, advMax:-1};
  for (const ev of events){
    const comp = ev.competitions[0], st = comp.status.type, state = st.state;
    const note = (comp.altGameNote || "").replace("FIFA World Cup, ", "");
    const L = koLevel(note);
    const parts = comp.competitors.map(c => ({r:resolve(c.team.displayName),
                                              raw:c.team.displayName, c}));
    for (const p of parts){
      if (!p.r) continue;
      const team = p.r[1];
      if (!(team in info)) continue;
      const opp = parts.find(x => x !== p);
      const oppName = opp ? (opp.r ? opp.r[1] : opp.raw) : "TBD";
      if (L >= 0) info[team].inKO = true;
      if (state === "pre" || state === "in"){
        const cand = {short:KO_SHORT[note] || note, opp:oppName, when:st.detail, date:ev.date};
        if (!info[team].next || ev.date < info[team].next.date) info[team].next = cand;
      } else if (state === "post" && L >= 0){
        if (p.c.advance || p.c.winner){
          info[team].advMax = Math.max(info[team].advMax, L);
          if (L === KO_ROUNDS.length - 1) info[team].champ = true;
        } else {
          const pens = (st.detail || "").toLowerCase().includes("pen");
          info[team].exit = {short:KO_SHORT[note] || note, by:oppName, pens};
        }
      }
    }
  }
  for (const t of teams){
    const o = info[t];
    if (o.champ){
      o.badge = "🏆 CHAMP"; o.cls = "champ"; o.sub = "Won the World Cup 🎉"; o.dead = false;
    } else if (o.exit){
      o.badge = "OUT"; o.cls = "out"; o.dead = true;
      o.sub = `Out — lost ${o.exit.short}${o.exit.by ? " to " + o.exit.by : ""}`
            + (o.exit.pens ? " (pens)" : "");
    } else if (!o.inKO){
      o.badge = "OUT"; o.cls = "out"; o.dead = true; o.sub = "Out — group stage";
    } else if (o.next){
      o.badge = o.next.short; o.cls = "live"; o.dead = false;
      o.sub = `Next: ${o.next.short} · vs ${o.next.opp} · ${o.next.when}`;
    } else {
      const nr = KO_ROUNDS[o.advMax + 1];
      o.badge = nr ? KO_SHORT[nr] : "ALIVE"; o.cls = "live"; o.dead = false;
      o.sub = "Through — next opponent TBD";
    }
  }
  return info;
}

// --- player picker: each visitor highlights their own teams (saved per device) ---
const sel = document.getElementById("me");
let me = localStorage.getItem("wc-me") || "";
for (const owner of Object.keys(OWNERS).sort()){
  const opt = document.createElement("option");
  opt.value = owner; opt.textContent = owner;
  sel.appendChild(opt);
}
sel.value = me;
sel.addEventListener("change", () => {
  me = sel.value;
  if (me) localStorage.setItem("wc-me", me); else localStorage.removeItem("wc-me");
  paint();
});

let lastPool = [];

// tap-to-toggle the next-match detail; remembered by team across the 60s refresh
const expanded = new Set();
document.getElementById("rows").addEventListener("click", (e) => {
  const tr = e.target.closest("tr");
  if (!tr || !tr.dataset.team) return;
  const t = tr.dataset.team;
  if (expanded.has(t)) expanded.delete(t); else expanded.add(t);
  tr.classList.toggle("expanded");
});

function paint(){
  const rows = document.getElementById("rows");
  rows.innerHTML = "";
  lastPool.forEach((s, i) => {
    const tr = document.createElement("tr");
    const x = s._info || {};
    tr.dataset.team = s.team;
    if (expanded.has(s.team)) tr.classList.add("expanded");
    if (me && s.owner === me) tr.classList.add("mine-row");
    if (x.dead) tr.classList.add("out");
    const badge = x.badge ? ` <span class="bdg ${x.cls}">${x.badge}</span>` : "";
    const caret = x.sub ? ` <span class="caret">›</span>` : "";
    const sub = x.sub ? `<div class="ssub ${x.dead ? "dead" : ""}">${x.sub}</div>` : "";
    tr.innerHTML =
      `<td>${i+1}</td><td class="team"><span class="tname">${s.team}</span>${badge}${caret}${sub}</td><td>${s.owner}</td>` +
      `<td>${s.p}</td><td>${s.w}-${s.d}-${s.l}</td>` +
      `<td>${fmtGD(s)}</td><td class="pts">${s.pts}</td>`;
    rows.appendChild(tr);
  });
  const mineEl = document.getElementById("mine");
  if (me){
    const chips = lastPool.filter(s => s.owner === me).map(s => {
      const x = s._info || {};
      const badge = x.badge ? `<span class="bdg ${x.cls}">${x.badge}</span>` : "";
      const sub = x.sub ? `<div class="ssub ${x.dead ? "dead" : ""}">${x.sub}</div>` : "";
      return `<span class="chip"><b>${s.team}</b> ${badge}${sub}</span>`;
    }).join("");
    mineEl.innerHTML = `<h2>${me}'s teams</h2><div class="chips">${chips}</div>`;
  } else {
    mineEl.innerHTML = `<div class="hint">👆 Pick your name above to highlight your teams.</div>`;
  }
}

function render(pool, live){
  lastPool = pool;
  paint();
  const t = new Date().toLocaleTimeString([], {hour:"numeric", minute:"2-digit"});
  const champ = pool.find(s => s._info && s._info.champ);
  const tag = champ ? ` · <span class="live">🏆 ${champ.team} (${champ.owner}) wins the $240 pot!</span>`
                    : (live ? ` · <span class="live">${live} live now</span>` : "");
  document.getElementById("status").innerHTML = `Updated ${t}${tag}`;
}

async function load(){
  try{
    const resp = await fetch(ENDPOINT, {cache:"no-store"});
    const data = await resp.json();
    const events = data.events || [];
    const stats = {};
    for (const [owner, teams] of Object.entries(OWNERS))
      for (const variants of teams)
        stats[variants[0]] = {team:variants[0], owner, p:0, w:0, d:0, l:0, gf:0, ga:0};
    let live = 0;
    for (const ev of events){
      const comp = ev.competitions[0];
      const state = comp.status.type.state;
      if (state === "in") live++;
      if (state !== "in" && state !== "post") continue;
      const got = [];
      let ok = true;
      for (const c of comp.competitors){
        const r = resolve(c.team.displayName);
        if (!r || c.score === "" || c.score == null){ ok = false; break; }
        got.push([r[1], parseInt(c.score, 10)]);
      }
      if (!ok || got.length !== 2) continue;
      const [[ta, sa], [tb, sb]] = got;
      for (const [team, gf, ga] of [[ta, sa, sb], [tb, sb, sa]]){
        const s = stats[team];
        s.p++; s.gf += gf; s.ga += ga;
        if (gf > ga) s.w++; else if (gf < ga) s.l++; else s.d++;
      }
    }
    const pool = Object.values(stats);
    for (const s of pool){ s.pts = 3*s.w + s.d; s.gd = s.gf - s.ga; }
    const info = buildInfo(events, Object.keys(stats));
    for (const s of pool) s._info = info[s.team];
    // Sort in tiers: teams still alive first, then teams knocked out in the
    // knockout rounds (reached the KO but lost), then teams that never made
    // the knockouts (group-stage exits). Within each tier, standard soccer
    // order: Points -> goal diff -> goals for.
    const tier = s => { const x = s._info || {}; if (!x.dead) return 0; return x.inKO ? 1 : 2; };
    pool.sort((a,b) => tier(a)-tier(b) || b.pts-a.pts || b.gd-a.gd || b.gf-a.gf || a.team.localeCompare(b.team));
    render(pool, live);
  } catch (e){
    document.getElementById("status").textContent =
      "Couldn't reach ESPN — will retry in 60s…";
  }
}
load();
setInterval(load, 60000);
</script>
</body>
</html>
"""

def build_html(path):
    """Write the self-contained live standings page (for GitHub Pages)."""
    html = (HTML_TEMPLATE
            .replace("%%OWNERS%%", json.dumps(OWNERS, ensure_ascii=False, indent=2))
            .replace("%%ENDPOINT%%", ENDPOINT))
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)


# ---- live bracket page ----------------------------------------------------
# Companion to index.html: the full knockout bracket, round by round. Same
# self-contained pattern — fetches ESPN in the browser, shares the OWNERS map
# and the per-device name pick (localStorage "wc-me") with the standings page,
# and highlights the chosen owner's teams in gold. Future rounds show ESPN's
# own placeholders ("Quarterfinal 1 Winner") until the bracket is seeded.
HTML_BRACKET = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>World Cup 2026 Bracket</title>
<link rel="icon" href="data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'%3E%3Ctext y='.9em' font-size='90'%3E%E2%9A%BD%3C/text%3E%3C/svg%3E">
<meta name="theme-color" content="#1f3864">
<style>
  *{box-sizing:border-box}
  body{margin:0;font-family:-apple-system,"Segoe UI",Roboto,Arial,sans-serif;
       background:#0f1626;color:#e8edf6;-webkit-text-size-adjust:100%}
  header{background:#1f3864;padding:14px 12px;text-align:center}
  header h1{margin:0;font-size:20px;letter-spacing:.5px}
  header .sub{font-size:12px;opacity:.85;margin-top:3px}
  .nav{display:flex;gap:16px;justify-content:center;background:#16203a;
       padding:9px 12px;border-bottom:1px solid #1c2740;font-size:13px}
  .nav a{color:#9fb0cc;text-decoration:none;font-weight:600}
  .nav a.active{color:#ffd86b}
  .controls{display:flex;align-items:center;justify-content:center;gap:10px;
            flex-wrap:wrap;padding:10px 12px;background:#16203a;
            border-bottom:1px solid #1c2740}
  .controls label{font-size:13px;color:#9fb0cc}
  .controls select{background:#0f1626;color:#e8edf6;border:1px solid #2a3a5e;
                   border-radius:8px;padding:7px 10px;font-size:15px}
  .controls .only{font-size:13px;color:#9fb0cc;display:flex;align-items:center;gap:5px}
  .status{font-size:12px;text-align:center;padding:8px;color:#9fb0cc}
  .status .live{color:#ff6b6b;font-weight:700}
  /* horizontal scroll of round columns — natural swipe on phones, full bracket on desktop */
  .scroll{overflow-x:auto;-webkit-overflow-scrolling:touch;padding:12px}
  .bracket{display:flex;gap:14px;align-items:flex-start;min-width:min-content}
  .col{flex:0 0 auto;width:212px}
  .col h3{margin:0 0 10px;font-size:12px;text-transform:uppercase;letter-spacing:.5px;
          color:#9fb0cc;text-align:center;position:sticky;top:0}
  .card{background:#16203a;border:1px solid #233152;border-radius:10px;
        padding:7px 9px;margin-bottom:10px}
  .card.minecard{border-color:#ffd86b;box-shadow:0 0 0 1px #ffd86b inset}
  .trow{display:flex;justify-content:space-between;align-items:center;gap:6px;
        padding:3px 0;font-size:14px}
  .trow+.trow{border-top:1px dashed #233152}
  .trow .tn{display:flex;flex-direction:column;line-height:1.15}
  .trow.win .tn{font-weight:700}
  .trow.win .sc{font-weight:700}
  .trow.ph .tn{color:#5d6b88;font-style:italic;font-size:12px}
  .trow.mineteam .tn{color:#ffd86b}
  .trow .own{font-size:10px;color:#7d8db0;font-weight:600;text-transform:uppercase;
             letter-spacing:.3px}
  .trow.mineteam .own{color:#e9c25a}
  .sc{min-width:18px;text-align:right;color:#cdd8ee}
  .cfoot{margin-top:5px;font-size:10px;color:#6b7a99;text-align:center}
  .cfoot.pens{color:#e0a85a;font-weight:600}
  .cfoot.livegame{color:#ff6b6b;font-weight:700}
  .empty{font-size:12px;color:#566685;text-align:center;padding:10px 0}
  /* desktop connector-line tree (used at >=900px; narrower screens get the columns above) */
  .thead{position:relative;height:22px;margin:0 auto 6px}
  .thead div{position:absolute;text-align:center;font-size:11px;text-transform:uppercase;
             letter-spacing:.5px;color:#9fb0cc;font-weight:600}
  .tree{position:relative;margin:0 auto}
  .tree .lines{position:absolute;top:0;left:0;pointer-events:none}
  .tree .lines path{fill:none;stroke:#2a3a5e;stroke-width:2}
  .tree .card{position:absolute;margin:0;height:66px;overflow:hidden;padding:5px 9px}
  .tree .card.dim{opacity:.28}
  .tree .trow{padding:2px 0;font-size:13px}
  .tree .trow .tn{flex-direction:row;align-items:baseline;gap:6px}
  .tree .own{font-size:9px}
  .tree .cfoot{margin-top:2px}
  footer{font-size:11px;text-align:center;color:#6b7a99;padding:14px 12px 24px;line-height:1.5}
</style>
</head>
<body>
<header>
  <h1>⚽ WORLD CUP 2026 BRACKET</h1>
  <div class="sub">Knockout rounds · winner-take-all · $240 pot</div>
</header>
<nav class="nav"><a href="index.html">← Standings</a><a class="active" href="bracket.html">Bracket</a></nav>
<div class="controls">
  <label for="me">⭐ My teams:</label>
  <select id="me"><option value="">— pick your name —</option></select>
  <label class="only"><input type="checkbox" id="onlyMine"> only my path</label>
</div>
<div class="status" id="status">Loading bracket…</div>
<div class="scroll"><div id="view"></div></div>
<footer>
  Live from ESPN (unofficial) · auto-refreshes every 60s · your teams in gold<br>
  Future rounds show ESPN's placeholders until the bracket is seeded.
</footer>
<script>
const OWNERS = %%OWNERS%%;
const ENDPOINT = "%%ENDPOINT%%";

function norm(s){
  return (s||"").normalize("NFKD").replace(/[̀-ͯ]/g,"")
                .toLowerCase().replace(/[^a-z0-9]/g,"");
}
const LUT = {};
for (const [owner, teams] of Object.entries(OWNERS))
  for (const variants of teams)
    for (const v of variants) LUT[norm(v)] = [owner, variants[0]];
function resolve(name){ return LUT[norm(name)] || null; }

const KO_ROUNDS = ["Round of 32","Round of 16","Quarterfinals","Semifinals","Final"];
function koLevel(note){
  for (let i = 0; i < KO_ROUNDS.length; i++)
    if (note.endsWith(KO_ROUNDS[i])) return i;
  return -1;
}

// --- name picker (shared with the standings page via localStorage) ---
const sel = document.getElementById("me");
const onlyEl = document.getElementById("onlyMine");
let me = localStorage.getItem("wc-me") || "";
for (const owner of Object.keys(OWNERS).sort()){
  const opt = document.createElement("option");
  opt.value = owner; opt.textContent = owner;
  sel.appendChild(opt);
}
sel.value = me;
onlyEl.disabled = !me;
sel.addEventListener("change", () => {
  me = sel.value;
  if (me) localStorage.setItem("wc-me", me); else localStorage.removeItem("wc-me");
  onlyEl.disabled = !me;
  if (!me) onlyEl.checked = false;
  paint();
});
onlyEl.addEventListener("change", paint);
window.addEventListener("resize", paint);

let lastEvents = [];

// Parse an unseeded slot ("Round of 32 5 Winner") -> [level, matchNumber].
function placeholderFeeder(raw){
  let m;
  if ((m = raw.match(/^Round of 32 (\d+) Winner/)))  return [0, +m[1]];
  if ((m = raw.match(/^Round of 16 (\d+) Winner/)))  return [1, +m[1]];
  if ((m = raw.match(/^Quarterfinal (\d+) Winner/)))  return [2, +m[1]];
  if ((m = raw.match(/^Semifinal (\d+) Winner/)))     return [3, +m[1]];
  return null;
}

// Build the knockout matches grouped by round, plus each match's two feeder
// matches and a vertical layout position. Feeders come from the placeholder
// names while a slot is unseeded, and from winner-matching once it's filled —
// so the (irregular) bracket topology reconstructs correctly at any stage.
function buildModel(){
  const byRound = [[], [], [], [], []];
  for (const ev of lastEvents){
    const comp = ev.competitions[0], st = comp.status.type;
    const note = (comp.altGameNote || "").replace("FIFA World Cup, ", "");
    const L = koLevel(note);
    if (L < 0) continue;
    const teams = comp.competitors.map(c => {
      const r = resolve(c.team.displayName);
      return {name: r ? r[1] : c.team.displayName, raw: c.team.displayName,
              owner: r ? r[0] : null, ph: !r, score: c.score, win: !!c.winner,
              mine: !!(me && r && r[0] === me)};
    });
    byRound[L].push({id:+ev.id, level:L, st, date:ev.date, teams,
                     mineHere: teams.some(t => t.mine)});
  }
  for (const r of byRound) r.sort((a, b) => a.id - b.id);
  const ok = byRound[0].length===16 && byRound[1].length===8 &&
             byRound[2].length===4 && byRound[3].length===2 && byRound[4].length===1;
  if (ok){
    const numMap = byRound.map(r => { const m = {}; r.forEach((mt,i) => m[i+1] = mt); return m; });
    for (let L = 1; L < 5; L++)
      for (const mt of byRound[L])
        mt.feeders = mt.teams.map(t => {
          if (t.ph){ const f = placeholderFeeder(t.raw); return f ? numMap[f[0]][f[1]] : null; }
          const nT = norm(t.name);
          return byRound[L-1].find(pm => pm.teams.some(pt => pt.win && norm(pt.name) === nT)) || null;
        });
    const leaves = [];
    (function visit(mt){
      if (mt.level === 0){ leaves.push(mt); return; }
      (mt.feeders || []).forEach(f => { if (f) visit(f); });
    })(byRound[4][0]);
    const GAP = 80;
    leaves.forEach((mt, i) => mt.center = i*GAP + GAP/2);
    for (let L = 1; L < 5; L++)
      for (const mt of byRound[L]){
        const cs = (mt.feeders || []).filter(Boolean).map(f => f.center);
        mt.center = cs.length ? cs.reduce((a,b) => a+b, 0)/cs.length : 0;
      }
    return {byRound, ok:true, height: 16*GAP + 8};
  }
  return {byRound, ok:false};
}

function cardHtml(mt){
  const state = mt.st.state, pens = (mt.st.detail || "").toLowerCase().includes("pen");
  const rows = mt.teams.map(t => {
    const cls = [t.ph ? "ph" : "", t.win ? "win" : "", t.mine ? "mineteam" : ""].join(" ");
    const own = t.owner ? `<span class="own">${t.owner}</span>` : "";
    const sc = state === "pre" ? "" : `<span class="sc">${t.score}</span>`;
    return `<div class="trow ${cls}"><span class="tn">${t.name}${own}</span>${sc}</div>`;
  }).join("");
  let foot, fcls = "";
  if (state === "post"){ foot = mt.st.detail || "Final"; if (pens) fcls = "pens"; }
  else if (state === "in"){ foot = "LIVE " + (mt.st.detail || ""); fcls = "livegame"; }
  else { foot = mt.st.detail || ""; }
  return rows + `<div class="cfoot ${fcls}">${foot}</div>`;
}

function renderColumns(model){
  const onlyMine = onlyEl.checked && me;
  let html = `<div class="bracket">`;
  model.byRound.forEach((matches, L) => {
    html += `<div class="col"><h3>${KO_ROUNDS[L]}</h3>`;
    let shown = 0;
    for (const mt of matches){
      if (onlyMine && !mt.mineHere) continue;
      shown++;
      html += `<div class="card${mt.mineHere ? " minecard" : ""}">${cardHtml(mt)}</div>`;
    }
    if (!shown) html += `<div class="empty">${onlyMine ? "— none —" : "(scheduled)"}</div>`;
    html += `</div>`;
  });
  document.getElementById("view").innerHTML = html + `</div>`;
}

function renderTree(model){
  const onlyMine = onlyEl.checked && me;
  const COLW = 212, CARDW = 188, CARDH = 66, W = 5*COLW;
  let hdr = `<div class="thead" style="width:${W}px">`;
  KO_ROUNDS.forEach((r, L) => hdr += `<div style="left:${L*COLW}px;width:${CARDW}px">${r}</div>`);
  hdr += `</div>`;
  let paths = "";
  for (let L = 1; L < 5; L++)
    for (const mt of model.byRound[L]){
      const px = L*COLW, py = mt.center;
      (mt.feeders || []).forEach(f => {
        if (!f) return;
        const cx = (L-1)*COLW + CARDW, cy = f.center, mx = (cx + px)/2;
        paths += `<path d="M${cx} ${cy} H${mx} V${py} H${px}" />`;
      });
    }
  let cards = "";
  for (let L = 0; L < 5; L++)
    for (const mt of model.byRound[L]){
      const dim = onlyMine && !mt.mineHere ? " dim" : "";
      cards += `<div class="card${mt.mineHere ? " minecard" : ""}${dim}" `
             + `style="left:${L*COLW}px;top:${mt.center - CARDH/2}px;width:${CARDW}px">`
             + `${cardHtml(mt)}</div>`;
    }
  document.getElementById("view").innerHTML = hdr +
    `<div class="tree" style="width:${W}px;height:${model.height}px">` +
    `<svg class="lines" width="${W}" height="${model.height}">${paths}</svg>${cards}</div>`;
}

function paint(){
  const model = buildModel();
  if (model.ok && window.innerWidth >= 900) renderTree(model);
  else renderColumns(model);
}

async function load(){
  try{
    const resp = await fetch(ENDPOINT, {cache:"no-store"});
    const data = await resp.json();
    lastEvents = data.events || [];
    let live = 0, champ = null;
    for (const ev of lastEvents){
      const comp = ev.competitions[0], st = comp.status.type;
      if (st.state === "in") live++;
      const note = (comp.altGameNote || "").replace("FIFA World Cup, ", "");
      if (st.state === "post" && koLevel(note) === KO_ROUNDS.length - 1){
        const w = comp.competitors.find(c => c.winner);
        if (w){ const r = resolve(w.team.displayName); champ = r ? `${r[1]} (${r[0]})` : w.team.displayName; }
      }
    }
    paint();
    const t = new Date().toLocaleTimeString([], {hour:"numeric", minute:"2-digit"});
    const tag = champ ? ` · <span class="live">🏆 ${champ} wins the pot!</span>`
                      : (live ? ` · <span class="live">${live} live now</span>` : "");
    document.getElementById("status").innerHTML = `Updated ${t}${tag}`;
  } catch (e){
    document.getElementById("status").textContent =
      "Couldn't reach ESPN — will retry in 60s…";
  }
}
load();
setInterval(load, 60000);
</script>
</body>
</html>
"""

def build_bracket(path):
    """Write the self-contained live bracket page (companion to index.html)."""
    html = (HTML_BRACKET
            .replace("%%OWNERS%%", json.dumps(OWNERS, ensure_ascii=False, indent=2))
            .replace("%%ENDPOINT%%", ENDPOINT))
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)

# ---- console summary ------------------------------------------------------
def print_summary(data, pool=None):
    print("\n=== Steve's World Cup teams ===")
    for t in MY_TEAMS:
        print(f"\n{t} (Group {GROUP[t]}):")
        if not data[t]:
            print("  (no fixtures found yet)")
        for m in data[t]:
            score = f"{m['gf']}-{m['ga']}" if m["gf"] is not None else "  - "
            tag = "[LIVE]" if m["state"] == "in" else ("[FT]" if m["state"] == "post" else "")
            print(f"  {m['opponent']:<22} {score:>5}  {tag:<6} {m['round'] or 'Group'}")
    if pool:
        print("\n=== Pool standings (Pts -> GD -> GF) ===")
        print(f"  {'#':>2}  {'Team':<22} {'Owner':<8} {'P':>2} {'W-D-L':>6} "
              f"{'GD':>4} {'Pts':>3}")
        for rank, s in enumerate(pool, start=1):
            star = "*" if s["owner"] == ME else " "
            wdl = f'{s["w"]}-{s["d"]}-{s["l"]}'
            gd = f'{s["gd"]:+d}' if s["p"] else "0"
            print(f"{star} {rank:>2}  {s['team']:<22} {s['owner']:<8} {s['p']:>2} "
                  f"{wdl:>6} {gd:>4} {s['pts']:>3}")

# ---- main -----------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=default_output_path())
    ap.add_argument("--fixture", default=None, help="local JSON file for offline testing")
    ap.add_argument("--html", nargs="?", const="index.html", default=None,
                    help="also write the live web pages (default: index.html, plus "
                         "bracket.html alongside it) for GitHub Pages; the pages "
                         "fetch ESPN themselves, so they don't depend on this run's data")
    a = ap.parse_args()
    try:
        events = fetch_events(a.fixture)
    except Exception as e:
        print(f"Fetch failed: {e}", file=sys.stderr); sys.exit(1)
    data = collect(events)
    pool = pool_table(events)
    build_xlsx(data, pool, a.out)
    print_summary(data, pool)
    print(f"\nWrote {a.out}")
    if a.html:
        build_html(a.html)
        print(f"Wrote {a.html}")
        bracket_path = os.path.join(os.path.dirname(a.html) or ".", "bracket.html")
        build_bracket(bracket_path)
        print(f"Wrote {bracket_path}")

if __name__ == "__main__":
    main()
